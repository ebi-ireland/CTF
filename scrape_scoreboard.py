"""
zerodays.events - 個人スコアランキング抽出
- .envからUSERNAME/PASSWORDを読み込みログイン
- スコアボードページ: Selenium (JS描画待ち) + Chrome visible mode
- 各チームページ: requests (静的HTML) → 高速
"""

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import time, os, re

BASE_URL = "https://zerodays.events"
SESSION  = requests.Session()
SESSION.headers.update({"User-Agent": "Mozilla/5.0"})

# ── .envの読み込み ────────────────────────────────────────────────────────────
def load_env():
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    env = {}
    with open(env_path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                key, _, val = line.partition("=")
                env[key.strip()] = val.strip().strip('"').strip("'")
    username = env.get("USERNAME") or env.get("username")
    password = env.get("PASSWORD") or env.get("password")
    if not username or not password:
        raise ValueError(".envにUSERNAMEとPASSWORDが見つかりません")
    return username, password

# ── Step 1: ログイン → スコアボード → チーム一覧取得 ─────────────────────────
def get_teams_via_selenium(username, password):
    opts = webdriver.ChromeOptions()
    # 見える状態で起動（headlessなし）
    opts.add_argument("--window-size=1280,900")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")

    print("[Chrome起動] ログインページを開きます...")
    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=opts
    )

    try:
        # ── ログイン ──────────────────────────────────────────────────────────
        driver.get(f"{BASE_URL}/login")
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "name"))
        )

        driver.find_element(By.ID, "name").send_keys(username)
        driver.find_element(By.ID, "password").send_keys(password)
        driver.find_element(By.ID, "_submit").click()

        print("[待機中] ログイン完了を確認しています...")
        # challengesページに遷移するまで待機
        WebDriverWait(driver, 15).until(
            EC.url_contains("/challenges")
        )
        print(f"[ログイン成功] 現在のURL: {driver.current_url}")

        # ── スコアボードへ移動 ────────────────────────────────────────────────
        print("[移動] スコアボードページへ...")
        driver.get(f"{BASE_URL}/scoreboard")

        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "tbody tr td a[href*='/teams/']"))
        )
        time.sleep(1.5)  # Alpine.js 描画待ち
        print(f"[完了] スコアボード読み込み完了")

        # ── チーム一覧取得 ────────────────────────────────────────────────────
        links = driver.find_elements(By.CSS_SELECTOR, "tbody tr td a[href*='/teams/']")
        teams = []
        seen  = set()
        for a in links:
            href = a.get_attribute("href")
            name = a.text.strip()
            if not href or not name:
                continue
            path = "/" + "/".join(href.split("/")[3:])  # /teams/130
            if path not in seen:
                seen.add(path)
                teams.append({"name": name, "href": path})

        # ── Seleniumのクッキーをrequestsセッションに移植 ─────────────────────
        for cookie in driver.get_cookies():
            SESSION.cookies.set(cookie["name"], cookie["value"])

        print(f"[完了] {len(teams)}チームを取得。Chromeを閉じます。")
        return teams

    finally:
        driver.quit()

# ── Step 2: requestsで各チームページからメンバー取得 ─────────────────────────
def get_members(team_name, team_href):
    url = f"{BASE_URL}{team_href}"
    try:
        resp = SESSION.get(url, timeout=10)
        resp.raise_for_status()
    except Exception as e:
        print(f"    ERROR: {e}")
        return []

    soup = BeautifulSoup(resp.text, "html.parser")
    members = []

    h3 = soup.find("h3", string=re.compile(r"Members", re.I))
    if not h3:
        return []
    table = h3.find_next("table")
    if not table:
        return []

    for row in table.select("tbody tr"):
        cols = row.find_all("td")
        if len(cols) < 2:
            continue
        a_tag  = cols[0].find("a")
        username = a_tag.get_text(strip=True) if a_tag else cols[0].get_text(strip=True)
        username = re.sub(r"\s*Captain\s*", "", username).strip()
        try:
            score = int(cols[1].get_text(strip=True))
        except ValueError:
            score = 0
        if username:
            members.append({"username": username, "score": score, "team": team_name})

    return members

# ── Step 3: Excel出力 ────────────────────────────────────────────────────────
def create_excel(all_players):
    all_players.sort(key=lambda x: x["score"], reverse=True)

    today    = datetime.now().strftime("%Y_%m_%d")
    filename = f"scoreboard_{today}.xlsx"
    candidates = [
        r"C:\Users\user\OneDrive\Desktop",
        os.path.expanduser("~/Desktop"),
        "/mnt/user-data/outputs",
        ".",
    ]
    out_dir  = next((d for d in candidates if os.path.exists(d)), ".")
    filepath = os.path.join(out_dir, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Player Rankings"

    HDR_FILL = PatternFill("solid", start_color="1F4E79")
    ALT_FILL = PatternFill("solid", start_color="D6E4F0")
    WHT_FILL = PatternFill("solid", start_color="FFFFFF")
    HDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    BODY_FONT = Font(name="Arial", size=10)
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT   = Alignment(horizontal="left",   vertical="center")
    thin   = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Rank", "User Name", "Score", "Team"]
    widths  = [8, 30, 12, 45]
    ws.row_dimensions[1].height = 24

    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = CENTER; cell.border = BORDER
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    for i, p in enumerate(all_players, 1):
        fill = ALT_FILL if i % 2 == 0 else WHT_FILL
        for c, (val, aln) in enumerate(
                zip([i, p["username"], p["score"], p["team"]], [CENTER, LEFT, CENTER, LEFT]), 1):
            cell = ws.cell(row=i+1, column=c, value=val)
            cell.font = BODY_FONT; cell.fill = fill
            cell.alignment = aln; cell.border = BORDER

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{len(all_players)+1}"
    wb.save(filepath)
    return filepath

# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    # .env読み込み
    print(".envを読み込み中...")
    username, password = load_env()
    print(f"ユーザー名: {username}")

    # Seleniumでログイン＆チーム一覧取得
    teams = get_teams_via_selenium(username, password)
    if not teams:
        print("チームが取得できませんでした。終了します。")
        return

    # requestsで各チームページを高速スクレイピング
    all_players = []
    for i, team in enumerate(teams, 1):
        print(f"  [{i:3d}/{len(teams)}] {team['name']}")
        members = get_members(team["name"], team["href"])
        print(f"           → {len(members)}人")
        all_players.extend(members)
        time.sleep(0.2)

    print(f"\n合計 {len(all_players)} 人分のデータを取得しました。")
    print("Excelを作成中...")
    filepath = create_excel(all_players)
    print(f"完了！ → {filepath}")

if __name__ == "__main__":
    main()